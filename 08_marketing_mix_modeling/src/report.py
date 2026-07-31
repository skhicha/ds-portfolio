"""
Automated Excel report generation for the MMM tool.

Builds a client-deliverable-style workbook (output/mmm_report.xlsx) with:
    1. "Summary"       - headline KPIs (R^2, MAE, MAPE, total sales, spend, ROI)
    2. "Channel Contribution" - contribution table + a native Excel bar chart
    3. "Model Coefficients"   - fitted Ridge coefficients per feature
    4. "Weekly Data"          - the underlying actual vs. fitted weekly series

Usage:
    python -m src.report                # fits the model on data/weekly_sales.csv
                                         # and writes output/mmm_report.xlsx
"""
from __future__ import annotations

import os
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.mmm import MMMResult, contribution_summary

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(size=16, bold=True, color="1F4E78")
SUBTITLE_FONT = Font(size=10, italic=True, color="595959")
KPI_LABEL_FONT = Font(bold=True)
KPI_VALUE_FONT = Font(size=14, bold=True, color="1F4E78")


def _style_header_row(ws, row_idx: int, n_cols: int) -> None:
    for col_idx in range(1, n_cols + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def _autofit_columns(ws, widths: dict) -> None:
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def _write_summary_sheet(wb: Workbook, result: MMMResult, contrib_summary: pd.DataFrame) -> None:
    ws = wb.active
    ws.title = "Summary"

    ws["B2"] = "Marketing Mix Modeling & Sales Forecasting - Summary Report"
    ws["B2"].font = TITLE_FONT
    ws["B3"] = f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M')} | " \
               f"{len(result.actuals)} weekly observations"
    ws["B3"].font = SUBTITLE_FONT

    total_sales = float(result.actuals.sum())
    total_spend = float(result.raw_spend.sum().sum())
    baseline_total = float(result.contributions["baseline"].sum())
    media_total = total_sales - baseline_total
    blended_roi = media_total / total_spend if total_spend else 0.0

    kpis = [
        ("Model R-squared", f"{result.r2:.3f}"),
        ("Mean Absolute Error (weekly sales)", f"{result.mae:,.0f}"),
        ("Mean Absolute % Error", f"{result.mape * 100:.2f}%"),
        ("Total Sales (period)", f"{total_sales:,.0f}"),
        ("Total Media + Promo Spend (period)", f"{total_spend:,.0f}"),
        ("Baseline Sales (non-media)", f"{baseline_total:,.0f}"),
        ("Media-Driven Sales", f"{media_total:,.0f}"),
        ("Blended ROI (media-driven sales / spend)", f"{blended_roi:.2f}x"),
    ]

    start_row = 5
    ws.cell(row=start_row, column=2, value="KPI").font = HEADER_FONT
    ws.cell(row=start_row, column=3, value="Value").font = HEADER_FONT
    _style_header_row(ws, start_row, 0)
    ws.cell(row=start_row, column=2).fill = HEADER_FILL
    ws.cell(row=start_row, column=3).fill = HEADER_FILL

    for i, (label, value) in enumerate(kpis, start=start_row + 1):
        ws.cell(row=i, column=2, value=label).font = KPI_LABEL_FONT
        ws.cell(row=i, column=3, value=value).font = KPI_VALUE_FONT

    ws.cell(row=start_row + len(kpis) + 2, column=2,
            value="Channel spend & carryover / saturation parameters used in the fit:").font = Font(bold=True)
    header_row = start_row + len(kpis) + 3
    headers = ["Channel", "Total Spend", "Adstock Decay", "Saturation Half-point (gamma)"]
    for j, h in enumerate(headers):
        ws.cell(row=header_row, column=2 + j, value=h)
    _style_header_row(ws, header_row, 0)
    for j in range(len(headers)):
        ws.cell(row=header_row, column=2 + j).fill = HEADER_FILL
        ws.cell(row=header_row, column=2 + j).font = HEADER_FONT

    for i, ch in enumerate(result.channels, start=header_row + 1):
        p = result.channel_params[ch]
        ws.cell(row=i, column=2, value=ch)
        ws.cell(row=i, column=3, value=float(result.raw_spend[ch].sum()))
        ws.cell(row=i, column=4, value=round(p.decay, 3))
        ws.cell(row=i, column=5, value=round(p.gamma, 2))

    _autofit_columns(ws, {"A": 2, "B": 40, "C": 20, "D": 18, "E": 30})


def _write_contribution_sheet(wb: Workbook, contrib_summary: pd.DataFrame) -> None:
    ws = wb.create_sheet("Channel Contribution")
    ws["B2"] = "Channel Contribution Decomposition"
    ws["B2"].font = TITLE_FONT

    header_row = 4
    cols = ["component", "total_contribution", "share_of_sales"]
    headers = ["Component", "Total Sales Contribution", "Share of Sales"]
    for j, h in enumerate(headers):
        ws.cell(row=header_row, column=2 + j, value=h)
    _style_header_row(ws, header_row, 0)
    for j in range(len(headers)):
        ws.cell(row=header_row, column=2 + j).fill = HEADER_FILL
        ws.cell(row=header_row, column=2 + j).font = HEADER_FONT

    for i, row in enumerate(contrib_summary.itertuples(index=False), start=header_row + 1):
        ws.cell(row=i, column=2, value=row.component)
        ws.cell(row=i, column=3, value=round(row.total_contribution, 2))
        pct_cell = ws.cell(row=i, column=4, value=row.share_of_sales)
        pct_cell.number_format = "0.0%"

    last_row = header_row + len(contrib_summary)
    _autofit_columns(ws, {"A": 2, "B": 24, "C": 26, "D": 16})

    chart = BarChart()
    chart.type = "col"
    chart.title = "Sales Contribution by Component"
    chart.y_axis.title = "Total Sales Contribution"
    chart.x_axis.title = "Component"
    data = Reference(ws, min_col=3, min_row=header_row, max_row=last_row)
    cats = Reference(ws, min_col=2, min_row=header_row + 1, max_row=last_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 20
    chart.height = 11
    ws.add_chart(chart, f"F{header_row}")


def _write_coefficients_sheet(wb: Workbook, result: MMMResult) -> None:
    ws = wb.create_sheet("Model Coefficients")
    ws["B2"] = "Fitted Ridge Regression Coefficients"
    ws["B2"].font = TITLE_FONT
    ws["B3"] = "Features are adstocked + saturated spend (channels) or trend/seasonality terms."
    ws["B3"].font = SUBTITLE_FONT

    header_row = 5
    headers = ["Feature", "Coefficient"]
    for j, h in enumerate(headers):
        ws.cell(row=header_row, column=2 + j, value=h)
    _style_header_row(ws, header_row, 0)
    for j in range(len(headers)):
        ws.cell(row=header_row, column=2 + j).fill = HEADER_FILL
        ws.cell(row=header_row, column=2 + j).font = HEADER_FONT

    ws.cell(row=header_row + 1, column=2, value="Intercept")
    ws.cell(row=header_row + 1, column=3, value=round(result.intercept, 4))

    row = header_row + 2
    for name, coef in result.coefficients.items():
        ws.cell(row=row, column=2, value=name)
        ws.cell(row=row, column=3, value=round(float(coef), 4))
        row += 1

    ws.cell(row=row + 1, column=2, value="Model quality").font = Font(bold=True)
    ws.cell(row=row + 2, column=2, value="R-squared")
    ws.cell(row=row + 2, column=3, value=round(result.r2, 4))
    ws.cell(row=row + 3, column=2, value="MAE")
    ws.cell(row=row + 3, column=3, value=round(result.mae, 2))
    ws.cell(row=row + 4, column=2, value="MAPE")
    ws.cell(row=row + 4, column=3, value=f"{result.mape * 100:.2f}%")

    _autofit_columns(ws, {"A": 2, "B": 22, "C": 16})


def _write_weekly_data_sheet(wb: Workbook, result: MMMResult) -> None:
    ws = wb.create_sheet("Weekly Data")
    ws["B2"] = "Actual vs. Fitted Weekly Sales"
    ws["B2"].font = TITLE_FONT

    header_row = 4
    headers = ["Week", "Actual Sales", "Fitted Sales"] + list(result.channels)
    for j, h in enumerate(headers):
        ws.cell(row=header_row, column=2 + j, value=h)
    _style_header_row(ws, header_row, 0)
    for j in range(len(headers)):
        ws.cell(row=header_row, column=2 + j).fill = HEADER_FILL
        ws.cell(row=header_row, column=2 + j).font = HEADER_FONT

    dates = pd.to_datetime(result.dates).dt.strftime("%Y-%m-%d").tolist()
    for i in range(len(result.actuals)):
        row = header_row + 1 + i
        ws.cell(row=row, column=2, value=dates[i])
        ws.cell(row=row, column=3, value=round(float(result.actuals[i]), 2))
        ws.cell(row=row, column=4, value=round(float(result.fitted_values[i]), 2))
        for j, ch in enumerate(result.channels):
            ws.cell(row=row, column=5 + j, value=round(float(result.raw_spend[ch].iloc[i]), 2))

    last_row = header_row + len(result.actuals)
    chart = LineChart()
    chart.title = "Actual vs. Fitted Weekly Sales"
    chart.y_axis.title = "Weekly Sales"
    chart.x_axis.title = "Week"
    data = Reference(ws, min_col=3, max_col=4, min_row=header_row, max_row=last_row)
    chart.add_data(data, titles_from_data=True)
    chart.width = 24
    chart.height = 11
    ws.add_chart(chart, f"{get_column_letter(6 + len(result.channels))}{header_row}")

    _autofit_columns(ws, {"A": 2, "B": 14, "C": 16, "D": 16, "E": 14, "F": 16, "G": 20})


def generate_excel_report(result: MMMResult, output_path: str) -> str:
    """Build the full workbook and save it to `output_path`. Returns the path."""
    summary = contribution_summary(result)

    wb = Workbook()
    _write_summary_sheet(wb, result, summary)
    _write_contribution_sheet(wb, summary)
    _write_coefficients_sheet(wb, result)
    _write_weekly_data_sheet(wb, result)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    return output_path


def main() -> None:
    from src.mmm import fit_mmm

    project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    data_path = os.path.join(project_root, "data", "weekly_sales.csv")
    output_path = os.path.join(project_root, "output", "mmm_report.xlsx")

    df = pd.read_csv(data_path)
    result = fit_mmm(df)
    path = generate_excel_report(result, output_path)
    print(f"R^2 = {result.r2:.4f} | MAE = {result.mae:,.0f} | MAPE = {result.mape * 100:.2f}%")
    print(f"Excel report written to {path}")


if __name__ == "__main__":
    main()
