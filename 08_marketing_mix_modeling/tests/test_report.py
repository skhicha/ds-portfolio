import os

import openpyxl
import pytest

from src.mmm import fit_mmm
from src.report import generate_excel_report

EXPECTED_SHEETS = ["Summary", "Channel Contribution", "Model Coefficients", "Weekly Data"]


@pytest.fixture(scope="module")
def report_path(tmp_path_factory, synthetic_df):
    result = fit_mmm(synthetic_df)
    out_dir = tmp_path_factory.mktemp("report_output")
    out_path = os.path.join(str(out_dir), "mmm_report_test.xlsx")
    generate_excel_report(result, out_path)
    return out_path


def test_report_file_is_created(report_path):
    assert os.path.exists(report_path)
    assert os.path.getsize(report_path) > 0


def test_report_is_valid_xlsx_workbook(report_path):
    wb = openpyxl.load_workbook(report_path)
    assert wb is not None


def test_report_contains_expected_sheets(report_path):
    wb = openpyxl.load_workbook(report_path)
    for sheet_name in EXPECTED_SHEETS:
        assert sheet_name in wb.sheetnames


def test_summary_sheet_has_r2_kpi(report_path):
    wb = openpyxl.load_workbook(report_path)
    ws = wb["Summary"]
    all_values = [cell.value for row in ws.iter_rows() for cell in row]
    assert any(v == "Model R-squared" for v in all_values)


def test_channel_contribution_sheet_has_expected_columns(report_path):
    wb = openpyxl.load_workbook(report_path)
    ws = wb["Channel Contribution"]
    header_values = [ws.cell(row=4, column=c).value for c in range(2, 5)]
    assert header_values == ["Component", "Total Sales Contribution", "Share of Sales"]


def test_channel_contribution_sheet_has_chart(report_path):
    wb = openpyxl.load_workbook(report_path)
    ws = wb["Channel Contribution"]
    assert len(ws._charts) >= 1


def test_model_coefficients_sheet_lists_channels(report_path):
    wb = openpyxl.load_workbook(report_path)
    ws = wb["Model Coefficients"]
    all_values = [cell.value for row in ws.iter_rows() for cell in row]
    assert "tv_spend" in all_values
    assert "digital_spend" in all_values
    assert "promotions_spend" in all_values
    assert "Intercept" in all_values


def test_weekly_data_sheet_row_count_matches_observations(report_path, synthetic_df):
    wb = openpyxl.load_workbook(report_path)
    ws = wb["Weekly Data"]
    # header at row 4, data starts row 5
    data_rows = [r for r in ws.iter_rows(min_row=5, min_col=2, max_col=2) if r[0].value is not None]
    assert len(data_rows) == len(synthetic_df)
