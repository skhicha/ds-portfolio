"""Excel KPI dashboard generator (openpyxl) for the optimization results.

Produces output/territory_allocation_report.xlsx with 5 sheets:
    1. Summary               - headline KPIs, optimized vs. baseline, % improvement
    2. Optimized Allocation  - rep -> territory allocation table
    3. Territory KPIs        - per-territory coverage / revenue detail
    4. Rep KPIs              - per-rep utilization / cost detail
    5. Baseline Comparison   - side-by-side optimized vs. status-quo baseline
"""
from __future__ import annotations

import os

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

from src.optimize import OptimizationResult, allocation_to_long_df, evaluate_allocation

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(bold=True, size=14, color="1F4E78")
SUBTITLE_FONT = Font(italic=True, size=10, color="595959")
KPI_LABEL_FONT = Font(bold=True, size=11)
GOOD_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
BAD_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")


def _write_dataframe(ws, df: pd.DataFrame, start_row: int = 1, start_col: int = 1,
                      number_formats: dict | None = None):
    """Write a DataFrame to a worksheet starting at (start_row, start_col),
    with a styled header row. Returns the row index after the table."""
    number_formats = number_formats or {}
    for j, col_name in enumerate(df.columns):
        cell = ws.cell(row=start_row, column=start_col + j, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for i, (_, row) in enumerate(df.iterrows()):
        for j, col_name in enumerate(df.columns):
            cell = ws.cell(row=start_row + 1 + i, column=start_col + j, value=row[col_name])
            if col_name in number_formats:
                cell.number_format = number_formats[col_name]

    # Auto-size columns (approximate).
    for j, col_name in enumerate(df.columns):
        col_letter = get_column_letter(start_col + j)
        max_len = max([len(str(col_name))] + [len(str(v)) for v in df[col_name].astype(str)])
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 40)

    ws.freeze_panes = ws.cell(row=start_row + 1, column=start_col)
    return start_row + 1 + len(df)


def _build_optimized_sheet(wb: Workbook, opt_result: OptimizationResult):
    ws = wb.create_sheet("Optimized Allocation")
    ws["A1"] = "Optimized Rep -> Territory Allocation"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "Continuous LP solution: allocated_fte is the fraction of a rep's capacity assigned to a territory."
    ws["A2"].font = SUBTITLE_FONT

    alloc_df = allocation_to_long_df(opt_result.x, opt_result.reps_df, opt_result.territories_df)
    _write_dataframe(ws, alloc_df, start_row=4, number_formats={
        "allocated_fte": "0.000",
        "effective_fte_supplied": "0.000",
        "cost_contribution": "$#,##0",
    })
    return alloc_df


def _build_territory_sheet(wb: Workbook, opt_result: OptimizationResult, metrics: dict):
    ws = wb.create_sheet("Territory KPIs")
    ws["A1"] = "Territory-Level KPIs (Optimized Solution)"
    ws["A1"].font = TITLE_FONT

    t_df = opt_result.territories_df.copy()
    t_df["allocated_fte_supplied"] = np.round(metrics["territory_supplied_fte"], 3)
    t_df["coverage_pct"] = np.round(metrics["coverage"] * 100, 1)
    t_df["achieved_revenue"] = np.round(metrics["achieved_revenue"], 2)
    t_df["meets_min_coverage"] = np.where(metrics["meets_min_coverage"], "Yes", "No")

    cols = ["territory_id", "territory_name", "region", "required_fte",
            "allocated_fte_supplied", "min_coverage_pct", "coverage_pct",
            "potential_revenue", "achieved_revenue", "meets_min_coverage"]
    t_df = t_df[cols]

    end_row = _write_dataframe(ws, t_df, start_row=3, number_formats={
        "required_fte": "0.00",
        "allocated_fte_supplied": "0.00",
        "min_coverage_pct": "0%",
        "coverage_pct": '0.0"%"',
        "potential_revenue": "$#,##0",
        "achieved_revenue": "$#,##0",
    })

    # Conditional formatting for the meets_min_coverage column.
    col_idx = cols.index("meets_min_coverage") + 1
    for row in range(4, end_row):
        cell = ws.cell(row=row, column=col_idx)
        cell.fill = GOOD_FILL if cell.value == "Yes" else BAD_FILL

    # Small bar chart: potential vs achieved revenue by territory.
    chart = BarChart()
    chart.title = "Potential vs Achieved Revenue by Territory"
    chart.y_axis.title = "Revenue ($)"
    chart.x_axis.title = "Territory"
    data = Reference(ws, min_col=cols.index("potential_revenue") + 1,
                      max_col=cols.index("achieved_revenue") + 1,
                      min_row=3, max_row=end_row - 1)
    cats = Reference(ws, min_col=cols.index("territory_id") + 1, min_row=4, max_row=end_row - 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width, chart.height = 24, 10
    ws.add_chart(chart, f"{get_column_letter(len(cols) + 2)}3")


def _build_rep_sheet(wb: Workbook, opt_result: OptimizationResult, metrics: dict, alloc_df: pd.DataFrame):
    ws = wb.create_sheet("Rep KPIs")
    ws["A1"] = "Rep-Level KPIs (Optimized Solution)"
    ws["A1"].font = TITLE_FONT

    r_df = opt_result.reps_df.copy()
    r_df["allocated_fte"] = np.round(metrics["rep_allocated_fte"], 3)
    r_df["utilization_pct"] = np.round(metrics["utilization"] * 100, 1)
    r_df["cost_incurred"] = np.round(metrics["rep_cost"], 2)
    if not alloc_df.empty:
        n_territories = alloc_df.groupby("rep_id")["territory_id"].nunique()
    else:
        n_territories = pd.Series(dtype=int)
    r_df["num_territories_assigned"] = r_df["rep_id"].map(n_territories).fillna(0).astype(int)

    cols = ["rep_id", "rep_name", "region", "skill_level", "experience_years",
            "productivity_multiplier", "max_capacity_fte", "allocated_fte",
            "utilization_pct", "annual_cost", "cost_incurred", "num_territories_assigned"]
    r_df = r_df[cols]

    _write_dataframe(ws, r_df, start_row=3, number_formats={
        "productivity_multiplier": "0.00",
        "max_capacity_fte": "0.00",
        "allocated_fte": "0.000",
        "utilization_pct": '0.0"%"',
        "annual_cost": "$#,##0",
        "cost_incurred": "$#,##0",
    })


def _build_baseline_comparison_sheet(wb: Workbook, opt_metrics: dict, baseline_metrics: dict, baseline_label: str):
    ws = wb.create_sheet("Baseline Comparison")
    ws["A1"] = f"Optimized LP vs. Baseline ({baseline_label})"
    ws["A1"].font = TITLE_FONT

    def pct_change(new, old):
        if old == 0:
            return float("nan")
        return (new - old) / abs(old) * 100.0

    rows = [
        ("Total Revenue ($)", opt_metrics["total_revenue"], baseline_metrics["total_revenue"]),
        ("Total Cost ($)", opt_metrics["total_cost"], baseline_metrics["total_cost"]),
        ("Net Revenue ($)", opt_metrics["net_revenue"], baseline_metrics["net_revenue"]),
        ("Avg Territory Coverage (%)", opt_metrics["avg_coverage"] * 100, baseline_metrics["avg_coverage"] * 100),
        ("Avg Rep Utilization (%)", opt_metrics["avg_utilization"] * 100, baseline_metrics["avg_utilization"] * 100),
        ("# Territories Meeting Min. Coverage", opt_metrics["num_territories_meeting_min"],
         baseline_metrics["num_territories_meeting_min"]),
        ("# Territories Fully Covered (100%)", opt_metrics["num_territories_fully_covered"],
         baseline_metrics["num_territories_fully_covered"]),
    ]

    comp_df = pd.DataFrame(rows, columns=["Metric", "Optimized (LP)", f"Baseline ({baseline_label})"])
    comp_df["% Change"] = [round(pct_change(o, b), 2) for _, o, b in rows]

    end_row = _write_dataframe(ws, comp_df, start_row=3, number_formats={
        "Optimized (LP)": "#,##0.00",
        f"Baseline ({baseline_label})": "#,##0.00",
        "% Change": "+0.0%;-0.0%",
    })
    # Fix % Change number format (values are already in percent units, not fractions).
    col_idx = list(comp_df.columns).index("% Change") + 1
    for row in range(4, end_row):
        cell = ws.cell(row=row, column=col_idx)
        cell.number_format = '+0.0"%";-0.0"%"'
        if isinstance(cell.value, (int, float)) and not np.isnan(cell.value):
            cell.fill = GOOD_FILL if cell.value >= 0 else BAD_FILL


def _build_summary_sheet(wb: Workbook, opt_metrics: dict, baseline_metrics: dict, baseline_label: str,
                          reps_df: pd.DataFrame, territories_df: pd.DataFrame, budget):
    ws = wb.create_sheet("Summary", 0)
    ws["A1"] = "Sales Force Optimization & Territory Allocation Model"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "LP-optimized rep-to-territory allocation vs. baseline manual allocation"
    ws["A2"].font = SUBTITLE_FONT

    net_improvement_pct = None
    if baseline_metrics["net_revenue"] != 0:
        net_improvement_pct = ((opt_metrics["net_revenue"] - baseline_metrics["net_revenue"])
                                / abs(baseline_metrics["net_revenue"])) * 100.0

    kpis = [
        ("Number of Reps", len(reps_df)),
        ("Number of Territories", len(territories_df)),
        ("Optional Budget Cap", f"${budget:,.0f}" if budget else "None"),
        ("", ""),
        ("Optimized Total Revenue", f"${opt_metrics['total_revenue']:,.2f}"),
        ("Optimized Total Cost", f"${opt_metrics['total_cost']:,.2f}"),
        ("Optimized Net Revenue", f"${opt_metrics['net_revenue']:,.2f}"),
        ("", ""),
        (f"Baseline ({baseline_label}) Total Revenue", f"${baseline_metrics['total_revenue']:,.2f}"),
        (f"Baseline ({baseline_label}) Total Cost", f"${baseline_metrics['total_cost']:,.2f}"),
        (f"Baseline ({baseline_label}) Net Revenue", f"${baseline_metrics['net_revenue']:,.2f}"),
        ("", ""),
        ("Net Revenue Improvement vs. Baseline",
         f"{net_improvement_pct:+.1f}%" if net_improvement_pct is not None else "N/A"),
        ("Avg Territory Coverage (Optimized)", f"{opt_metrics['avg_coverage'] * 100:.1f}%"),
        ("Avg Rep Utilization (Optimized)", f"{opt_metrics['avg_utilization'] * 100:.1f}%"),
        ("Territories Meeting Min. Coverage (Optimized)",
         f"{opt_metrics['num_territories_meeting_min']} / {len(territories_df)}"),
        ("Territories Meeting Min. Coverage (Baseline)",
         f"{baseline_metrics['num_territories_meeting_min']} / {len(territories_df)}"),
    ]

    row = 4
    for label, value in kpis:
        ws.cell(row=row, column=1, value=label).font = KPI_LABEL_FONT
        ws.cell(row=row, column=2, value=value)
        row += 1

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 24


def generate_report(opt_result: OptimizationResult, baseline_x: np.ndarray, output_path: str,
                     baseline_label: str = "Status Quo"):
    """Build the full Excel workbook and save it to output_path."""
    reps_df = opt_result.reps_df
    territories_df = opt_result.territories_df

    opt_metrics = evaluate_allocation(opt_result.x, reps_df, territories_df)
    baseline_metrics = evaluate_allocation(baseline_x, reps_df, territories_df)

    wb = Workbook()
    # Remove the default sheet; we create our own in the desired order.
    default_sheet = wb.active
    wb.remove(default_sheet)

    alloc_df = _build_optimized_sheet(wb, opt_result)
    _build_territory_sheet(wb, opt_result, opt_metrics)
    _build_rep_sheet(wb, opt_result, opt_metrics, alloc_df)
    _build_baseline_comparison_sheet(wb, opt_metrics, baseline_metrics, baseline_label)
    _build_summary_sheet(wb, opt_metrics, baseline_metrics, baseline_label, reps_df, territories_df, opt_result.budget)

    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    wb.save(output_path)
    return output_path


def main():
    import argparse
    from src.optimize import load_data, solve_allocation
    from src.baseline import status_quo_allocation

    parser = argparse.ArgumentParser(description="Generate the Excel KPI dashboard report.")
    parser.add_argument("--reps", type=str, default=None)
    parser.add_argument("--territories", type=str, default=None)
    parser.add_argument("--budget", type=float, default=None)
    parser.add_argument("--out", type=str, default=None)
    args = parser.parse_args()

    reps_df, territories_df = load_data(args.reps, args.territories)
    opt_result = solve_allocation(reps_df, territories_df, budget=args.budget)
    if not opt_result.success:
        raise SystemExit(f"LP did not solve successfully: {opt_result.scipy_message}")

    baseline_x = status_quo_allocation(reps_df, territories_df)

    out_path = args.out or os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "output", "territory_allocation_report.xlsx"
    )
    path = generate_report(opt_result, baseline_x, out_path)
    print(f"Report written to {path}")


if __name__ == "__main__":
    main()
