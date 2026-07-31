import os

import openpyxl
import pytest

from src.optimize import solve_allocation
from src.baseline import status_quo_allocation
from src.report import generate_report

EXPECTED_SHEETS = [
    "Summary",
    "Optimized Allocation",
    "Territory KPIs",
    "Rep KPIs",
    "Baseline Comparison",
]


@pytest.fixture(scope="module")
def report_path(tmp_path_factory, full_dataset):
    reps_df, territories_df = full_dataset
    result = solve_allocation(reps_df, territories_df)
    assert result.success

    baseline_x = status_quo_allocation(reps_df, territories_df)
    out_dir = tmp_path_factory.mktemp("report_out")
    out_path = str(out_dir / "test_report.xlsx")
    generate_report(result, baseline_x, out_path)
    return out_path


def test_report_file_created(report_path):
    assert os.path.exists(report_path)
    assert os.path.getsize(report_path) > 0


def test_report_has_expected_sheets(report_path):
    wb = openpyxl.load_workbook(report_path)
    assert set(wb.sheetnames) == set(EXPECTED_SHEETS)
    # Summary should be first for readability.
    assert wb.sheetnames[0] == "Summary"


def test_optimized_allocation_sheet_columns(report_path):
    wb = openpyxl.load_workbook(report_path)
    ws = wb["Optimized Allocation"]
    header_row = [c.value for c in ws[4]]
    expected_cols = {"rep_id", "rep_name", "territory_id", "territory_name",
                     "allocated_fte", "effective_fte_supplied", "cost_contribution"}
    assert expected_cols.issubset(set(header_row))


def test_territory_kpi_sheet_columns(report_path):
    wb = openpyxl.load_workbook(report_path)
    ws = wb["Territory KPIs"]
    header_row = [c.value for c in ws[3]]
    expected_cols = {"territory_id", "territory_name", "required_fte",
                     "coverage_pct", "potential_revenue", "achieved_revenue",
                     "meets_min_coverage"}
    assert expected_cols.issubset(set(header_row))


def test_rep_kpi_sheet_columns(report_path):
    wb = openpyxl.load_workbook(report_path)
    ws = wb["Rep KPIs"]
    header_row = [c.value for c in ws[3]]
    expected_cols = {"rep_id", "rep_name", "utilization_pct", "cost_incurred",
                     "num_territories_assigned"}
    assert expected_cols.issubset(set(header_row))


def test_summary_sheet_has_kpi_values(report_path):
    wb = openpyxl.load_workbook(report_path)
    ws = wb["Summary"]
    labels = [ws.cell(row=r, column=1).value for r in range(1, 25)]
    assert "Optimized Net Revenue" in labels
    assert "Net Revenue Improvement vs. Baseline" in labels


def test_baseline_comparison_sheet_has_metrics(report_path):
    wb = openpyxl.load_workbook(report_path)
    ws = wb["Baseline Comparison"]
    header_row = [c.value for c in ws[3]]
    assert "Metric" in header_row
    assert "Optimized (LP)" in header_row
    metric_values = [ws.cell(row=r, column=1).value for r in range(4, 12)]
    assert "Total Revenue ($)" in metric_values
    assert "Net Revenue ($)" in metric_values
